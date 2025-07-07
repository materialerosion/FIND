import os
import pandas as pd
import datetime
from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from models.formula import db, Formula, Ingredient, FormulaIngredient, IngredientAlias
from sqlalchemy import and_, or_, func, distinct
from werkzeug.utils import secure_filename
import json
from io import BytesIO
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from xhtml2pdf import pisa

app = Flask(__name__)
CORS(app, origins=[
    "https://victorious-field-0dec9f70f.6.azurestaticapps.net", 
    "https://victorious-field-0dec9f70f.6.azurestaticapps.net/", 
    "http://localhost:3000"
])
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///formulas.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload size

# Ensure uploads directory exists
BACKUPS_DIR = os.path.join(os.path.dirname(__file__), 'backups')
os.makedirs(BACKUPS_DIR, exist_ok=True)
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db.init_app(app)

def initialize_database_from_excel():
    """Load the database from a predefined Excel file on server startup"""
    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], "Synaps Full 2025 Q1.xlsx")
    
    if not os.path.exists(excel_path):
        print(f"Warning: Default database file not found at {excel_path}")
        return False
    
    try:
        print(f"Loading database from {excel_path}...")
        df = pd.read_excel(excel_path)
        
        print(f"Successfully read Excel file with {len(df)} rows and {len(df.columns)} columns")
        
        # Backup existing aliases before clearing data
        print("Backing up existing aliases...")
        aliases_backup = []
        for alias in IngredientAlias.query.all():
            ingredient = Ingredient.query.get(alias.ingredient_id)
            if ingredient:
                aliases_backup.append({
                    'ingredient_name': ingredient.name,
                    'ingredient_number': ingredient.fing_item_number,
                    'alias': alias.alias
                })
        
        # Create an automatic backup if there are aliases
        if aliases_backup:
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_filename = f"aliases_backup_{timestamp}.json"
            backup_path = os.path.join(BACKUPS_DIR, backup_filename)
            
            # Ensure backup directory exists
            os.makedirs(BACKUPS_DIR, exist_ok=True)
            
            # Write the backup to a file
            with open(backup_path, 'w') as f:
                json.dump(aliases_backup, f, indent=2)
            print(f"Created automatic backup: {backup_filename}")
        
        # Clear existing data
        print("Clearing existing database records...")
        db.session.query(FormulaIngredient).delete()
        db.session.query(Formula).delete()
        db.session.query(IngredientAlias).delete()  # Delete aliases
        db.session.query(Ingredient).delete()
        db.session.commit()
        
        # Process the data
        ingredients_dict = {}
        formulas_dict = {}
        formula_ingredients_created = 0
        errors = 0
        
        # Process each row in the Excel file
        print("Processing Excel data...")
        for index, row in df.iterrows():
            if index % 5000 == 0:
                print(f"Processing row {index}/{len(df)}...")
            
            try:
                # Process ingredient if it doesn't exist
                fing_item_number = str(row['FING_ITEM_NUMBER'])
                if fing_item_number not in ingredients_dict:
                    ingredient = Ingredient(
                        fing_item_number=fing_item_number,
                        name=str(row['FING_DESCRIPTION']),
                        description=str(row['FING_DESCRIPTION']),
                        description_expanded=str(row['Ingredient Description Expanded'])
                    )
                    db.session.add(ingredient)
                    db.session.flush()  # Get the ID without committing
                    ingredients_dict[fing_item_number] = ingredient.id
                
                # Process formula if it doesn't exist
                object_number = str(row['OBJECT_NUMBER'])
                if object_number not in formulas_dict:
                    formula = Formula(
                        object_number=object_number,
                        formulation_name=str(row['FORMULATION_NAME']),
                        lifecycle_phase=str(row['LIFECYCLE_PHASE']),
                        formula_brand=str(row['FORMULA_BRAND']),
                        sbu_category=str(row['SBU_CATEGORY']),
                        dossier_type=str(row['DOSSIER_TYPE']),
                        regulatory_comments=str(row['REGULATORY_COMMENTS']) if 'REGULATORY_COMMENTS' in row and not pd.isna(row['REGULATORY_COMMENTS']) else '',
                        general_comments=str(row['GENERAL_COMMENTS']) if 'GENERAL_COMMENTS' in row and not pd.isna(row['GENERAL_COMMENTS']) else '',
                        production_sites=str(row['PRODUCTION_SITES_AVAILABLE']) if 'PRODUCTION_SITES_AVAILABLE' in row and not pd.isna(row['PRODUCTION_SITES_AVAILABLE']) else '',
                        predecessor_formulation_number=str(row['PREDECESSORFORMULATIONNUMBER']) if 'PREDECESSORFORMULATIONNUMBER' in row and not pd.isna(row['PREDECESSORFORMULATIONNUMBER']) else '',
                        successor_formulation_number=str(row['SUCCESSORFORMULATIONNUMBER']) if 'SUCCESSORFORMULATIONNUMBER' in row and not pd.isna(row['SUCCESSORFORMULATIONNUMBER']) else ''
                    )
                    db.session.add(formula)
                    db.session.flush()  # Get the ID without committing
                    formulas_dict[object_number] = formula.id
                
                # Create formula ingredient relationship
                # Handle NaN values in the amount field
                amount = 0.0  # Default value
                if 'FING_QUANTITY' in row:
                    if pd.notna(row['FING_QUANTITY']):
                        try:
                            # Try to convert to float
                            amount = float(row['FING_QUANTITY'])
                        except (ValueError, TypeError):
                            # If conversion fails, use the default value
                            if isinstance(row['FING_QUANTITY'], str) and row['FING_QUANTITY'].strip().upper() == 'Q.S.':
                                # Special case for "Q.S." (Quantum Satis - as much as needed)
                                amount = 0.0  # Use 0 as a placeholder
                                print(f"Row {index}: Q.S. value found, using 0.0 as placeholder")
                            else:
                                print(f"Row {index}: Invalid amount value: {row['FING_QUANTITY']}, using 0.0")
                
                unit = str(row['FING_UNIT']) if pd.notna(row['FING_UNIT']) else ''
                
                formula_ingredient = FormulaIngredient(
                    formula_id=formulas_dict[object_number],
                    ingredient_id=ingredients_dict[fing_item_number],
                    amount=amount,
                    unit=unit
                )
                db.session.add(formula_ingredient)
                formula_ingredients_created += 1

            

            except Exception as e:
                errors += 1
                print(f"Error processing row {index}: {str(e)}")
                if errors >= 100:  # Limit the number of errors before giving up
                    raise Exception(f"Too many errors ({errors}) during import. Last error: {str(e)}")
                continue
            
            # Commit in batches to avoid memory issues
            if index % 1000 == 0:
                try:
                    db.session.commit()
                    print(f"Committed batch at row {index}")
                except Exception as e:
                    db.session.rollback()
                    print(f"Error committing batch at row {index}: {str(e)}")
                    # Continue with the next batch
        
                # Final commit
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"Error during final commit: {str(e)}")
            return False
        
        # Restore aliases from the latest backup
        print("Restoring aliases from latest backup...")
        latest_backup = get_latest_backup()
        if latest_backup:
            aliases_restored = restore_aliases_from_backup(latest_backup)
            print(f"Restored {aliases_restored} aliases from backup")
        else:
            print("No backup found to restore aliases from")
        
        print(f"Successfully loaded database from Excel:")
        print(f"  - {len(ingredients_dict)} ingredients")
        print(f"  - {len(formulas_dict)} formulas")
        print(f"  - {formula_ingredients_created} formula-ingredient relationships")
        print(f"  - {errors} errors encountered and skipped")
        return True
        
    except Exception as e:
        db.session.rollback()
        print(f"Error loading database from Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def get_latest_backup():
    """Get the latest aliases backup from the server"""
    try:
        backups = []
        # List all JSON files in the backups directory
        if os.path.exists(BACKUPS_DIR):
            for filename in os.listdir(BACKUPS_DIR):
                if filename.endswith('.json') and filename.startswith('aliases_backup_'):
                    file_path = os.path.join(BACKUPS_DIR, filename)
                    file_stats = os.stat(file_path)
                    
                    backups.append({
                        'path': file_path,
                        'mtime': file_stats.st_mtime
                    })
            
            # Sort backups by modification time (newest first)
            if backups:
                backups.sort(key=lambda x: x['mtime'], reverse=True)
                latest_backup = backups[0]['path']
                
                # Load and return the backup data
                with open(latest_backup, 'r') as f:
                    return json.load(f)
    except Exception as e:
        print(f"Error getting latest backup: {str(e)}")
    
    return None

def restore_aliases_from_backup(aliases_data):
    """Restore aliases from backup data"""
    if not aliases_data:
        return 0
        
    aliases_restored = 0
    
    for alias_data in aliases_data:
        # Find the ingredient by name or number
        ingredient = Ingredient.query.filter(
            (Ingredient.name == alias_data['ingredient_name']) | 
            (Ingredient.fing_item_number == alias_data['ingredient_number'])
        ).first()
        
        if ingredient:
            # Check if this alias already exists for the ingredient
            existing_alias = IngredientAlias.query.filter_by(
                ingredient_id=ingredient.id,
                alias=alias_data['alias']
            ).first()
            
            if not existing_alias:
                new_alias = IngredientAlias(
                    alias=alias_data['alias'],
                    ingredient_id=ingredient.id
                )
                db.session.add(new_alias)
                aliases_restored += 1
    
    if aliases_restored > 0:
        db.session.commit()
    
    return aliases_restored

@app.route('/api/formulas', methods=['GET'])
def get_formulas():
    try:
        # Get pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 12, type=int)
        
        # Get filter parameters
        brand = request.args.get('brand', '')
        category = request.args.get('category', '')
        lifecycle_phase = request.args.get('lifecycle_phase', '')
        
        # Limit maximum per_page to avoid overwhelming responses
        per_page = min(per_page, 100)
        
        print(f"Fetching formulas (page {page}, per_page {per_page})...")
        
        # Build query with filters
        query = Formula.query
        
        if brand:
            query = query.filter(Formula.formula_brand == brand)
        
        if category:
            query = query.filter(Formula.sbu_category == category)
        
        if lifecycle_phase:
            query = query.filter(Formula.lifecycle_phase == lifecycle_phase)
        
        # Get paginated formulas
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        formulas = pagination.items
        
        print(f"Found {len(formulas)} formulas for this page")
        
        result = []
        for formula in formulas:
            formula_data = {
                'id': formula.id,
                'object_number': formula.object_number,
                'formulation_name': formula.formulation_name,
                'lifecycle_phase': formula.lifecycle_phase,
                'formula_brand': formula.formula_brand,
                'sbu_category': formula.sbu_category,
                'ingredients': []
            }
            
            for formula_ingredient in formula.ingredients:
                ingredient = Ingredient.query.get(formula_ingredient.ingredient_id)
                formula_data['ingredients'].append({
                    'id': ingredient.id,
                    'name': ingredient.name,
                    'fing_item_number': ingredient.fing_item_number,
                    'amount': formula_ingredient.amount,
                    'unit': formula_ingredient.unit
                })
            
            result.append(formula_data)
        
        # Include pagination metadata
        pagination_data = {
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': page,
            'per_page': per_page,
            'has_next': pagination.has_next,
            'has_prev': pagination.has_prev
        }
        
        print(f"Returning {len(result)} formulas with pagination data")
        return jsonify({
            'formulas': result,
            'pagination': pagination_data
        })
    except Exception as e:
        print(f"Error in get_formulas: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/api/formulas/<object_number>', methods=['GET'])
def get_formula(object_number):
    formula = Formula.query.filter_by(object_number=object_number).first_or_404()
    
    formula_data = {
        'id': formula.id,
        'object_number': formula.object_number,
        'formulation_name': formula.formulation_name,
        'lifecycle_phase': formula.lifecycle_phase,
        'formula_brand': formula.formula_brand,
        'sbu_category': formula.sbu_category,
        'dossier_type': formula.dossier_type,
        'regulatory_comments': formula.regulatory_comments,
        'general_comments': formula.general_comments,
        'production_sites': formula.production_sites,
        'predecessor_formulation_number': formula.predecessor_formulation_number,
        'successor_formulation_number': formula.successor_formulation_number,
        'ingredients': []
    }
    
    for formula_ingredient in formula.ingredients:
        ingredient = Ingredient.query.get(formula_ingredient.ingredient_id)
        formula_data['ingredients'].append({
            'id': ingredient.id,
            'name': ingredient.name,
            'fing_item_number': ingredient.fing_item_number,
            'description': ingredient.description,
            'description_expanded': ingredient.description_expanded,
            'amount': formula_ingredient.amount,
            'unit': formula_ingredient.unit
        })
    
    return jsonify(formula_data)

@app.route('/api/formulas/search', methods=['GET'])
def search_formulas():
    try:
        # Get pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 12, type=int)
        
        # Limit maximum per_page to avoid overwhelming responses
        per_page = min(per_page, 400)
        
        # Start with a base query
        query = Formula.query
        
        # Check for multiple ingredients with coupled amount ranges
        ingredient_filters = []
        i = 1
        while True:
            ingredient_param = request.args.get(f'ingredient{i}')
            if not ingredient_param:
                # If we have ingredient1 but not ingredient2, check for the legacy 'ingredient' param
                if i == 1:
                    ingredient_param = request.args.get('ingredient')
                    if ingredient_param:
                        min_amount = request.args.get('min_amount')
                        max_amount = request.args.get('max_amount')
                        ingredient_filters.append({
                            'name': ingredient_param,
                            'min_amount': min_amount,
                            'max_amount': max_amount
                        })
                break
            
            min_amount = request.args.get(f'min_amount{i}')
            max_amount = request.args.get(f'max_amount{i}')
            
            ingredient_filters.append({
                'name': ingredient_param,
                'min_amount': min_amount,
                'max_amount': max_amount
            })
            i += 1
        
        # If we have ingredient filters, apply them
        if ingredient_filters:
            # For each ingredient, we need to find formulas that contain it with the specified amount range
            formula_ids_with_all_ingredients = None
            
            for ingredient_filter in ingredient_filters:
                # Find ingredients matching the name OR any of their aliases
                ingredient_name = ingredient_filter['name']
                
                # Find all ingredients that match by name or alias
                matching_ingredients = db.session.query(Ingredient.id).distinct().\
                    outerjoin(IngredientAlias, Ingredient.id == IngredientAlias.ingredient_id).\
                    filter(
                        db.or_(
                            Ingredient.name.ilike(f'%{ingredient_name}%'),
                            IngredientAlias.alias.ilike(f'%{ingredient_name}%')
                        )
                    ).all()
                
                matching_ingredient_ids = [r[0] for r in matching_ingredients]
                
                if not matching_ingredient_ids:
                    # No ingredients match this filter, so no formulas will match
                    return jsonify({
                        'formulas': [],
                        'pagination': {
                            'total': 0,
                            'pages': 0,
                            'current_page': page,
                            'per_page': per_page,
                            'has_next': False,
                            'has_prev': False
                        }
                    })
                
                # Find formulas containing any of these ingredients
                subquery = db.session.query(FormulaIngredient.formula_id).\
                    filter(FormulaIngredient.ingredient_id.in_(matching_ingredient_ids))
                
                # Apply amount filters for this specific ingredient
                if ingredient_filter['min_amount']:
                    subquery = subquery.filter(FormulaIngredient.amount >= float(ingredient_filter['min_amount']))
                
                if ingredient_filter['max_amount']:
                    subquery = subquery.filter(FormulaIngredient.amount <= float(ingredient_filter['max_amount']))
                
                # Get distinct formula IDs that match this ingredient filter
                formula_ids_with_ingredient = [r[0] for r in subquery.distinct().all()]
                
                # Intersect with previous results
                if formula_ids_with_all_ingredients is None:
                    formula_ids_with_all_ingredients = set(formula_ids_with_ingredient)
                else:
                    formula_ids_with_all_ingredients &= set(formula_ids_with_ingredient)
                
                # If at any point we have no matching formulas, we can stop
                if not formula_ids_with_all_ingredients:
                    break
            
            # Apply the formula ID filter to our main query
            if formula_ids_with_all_ingredients:
                query = query.filter(Formula.id.in_(formula_ids_with_all_ingredients))
            else:
                # No formulas match all ingredients
                return jsonify({
                    'formulas': [],
                    'pagination': {
                        'total': 0,
                        'pages': 0,
                        'current_page': page,
                        'per_page': per_page,
                        'has_next': False,
                        'has_prev': False
                    }
                })
        
        # Check for ingredient exclusions
        exclude_ingredient_filters = []
        i = 1
        while True:
            exclude_param = request.args.get(f'exclude_ingredient{i}')
            if not exclude_param:
                break
            exclude_ingredient_filters.append(exclude_param)
            i += 1
        
        # If we have exclusions, find all formulas that contain any of the excluded ingredients
        if exclude_ingredient_filters:
            exclude_ingredient_ids = set()
            for exclude_name in exclude_ingredient_filters:
                matching_ingredients = db.session.query(Ingredient.id).distinct().\
                    outerjoin(IngredientAlias, Ingredient.id == IngredientAlias.ingredient_id).\
                    filter(
                        db.or_(
                            Ingredient.name.ilike(f'%{exclude_name}%'),
                            IngredientAlias.alias.ilike(f'%{exclude_name}%')
                        )
                    ).all()
                exclude_ingredient_ids.update([r[0] for r in matching_ingredients])
            if exclude_ingredient_ids:
                # Find all formula IDs that contain any of these ingredients
                exclude_formula_ids = db.session.query(FormulaIngredient.formula_id).\
                    filter(FormulaIngredient.ingredient_id.in_(exclude_ingredient_ids)).distinct().all()
                exclude_formula_ids = set([r[0] for r in exclude_formula_ids])
                if exclude_formula_ids:
                    query = query.filter(~Formula.id.in_(exclude_formula_ids))
        
        # Apply other filters
        brand = request.args.get('brand')
        if brand:
            query = query.filter(Formula.formula_brand == brand)
        
        category = request.args.get('category')
        if category:
            query = query.filter(Formula.sbu_category == category)
        
        lifecycle_phase = request.args.get('lifecycle_phase')
        if lifecycle_phase:
            query = query.filter(Formula.lifecycle_phase == lifecycle_phase)
        
        # Filter by production site (partial, case-insensitive match)
        production_site = request.args.get('production_site')
        if production_site:
            query = query.filter(Formula.production_sites.ilike(f'%{production_site}%'))
        
        # Make sure results are distinct
        query = query.distinct()
        
        # Count total results for pagination
        total_count = query.count()
        
        # Apply pagination
        formulas = query.paginate(page=page, per_page=per_page, error_out=False).items
        
        # Prepare result
        result = []
        for formula in formulas:
            formula_data = {
                'id': formula.id,
                'object_number': formula.object_number,
                'formulation_name': formula.formulation_name,
                'lifecycle_phase': formula.lifecycle_phase,
                'formula_brand': formula.formula_brand,
                'sbu_category': formula.sbu_category,
                'ingredients': []
            }
            
            for formula_ingredient in formula.ingredients:
                ingredient = Ingredient.query.get(formula_ingredient.ingredient_id)
                formula_data['ingredients'].append({
                    'id': ingredient.id,
                    'name': ingredient.name,
                    'fing_item_number': ingredient.fing_item_number,
                    'amount': formula_ingredient.amount,
                    'unit': formula_ingredient.unit
                })
            
            result.append(formula_data)
        
        # Calculate pagination metadata
        total_pages = (total_count + per_page - 1) // per_page  # Ceiling division
        
        pagination_data = {
            'total': total_count,
            'pages': total_pages,
            'current_page': page,
            'per_page': per_page,
            'has_next': page < total_pages,
            'has_prev': page > 1
        }
        
        return jsonify({
            'formulas': result,
            'pagination': pagination_data
        })
        
    except Exception as e:
        print(f"Error in search_formulas: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/api/aliases/export', methods=['GET'])
def export_aliases():
    try:
        # Query all aliases with their ingredient information
        aliases_data = []
        for alias in IngredientAlias.query.all():
            ingredient = Ingredient.query.get(alias.ingredient_id)
            if ingredient:
                aliases_data.append({
                    'Ingredient Name': ingredient.name,
                    'Alias': alias.alias
                })
        
        # Create a DataFrame from the aliases data
        df = pd.DataFrame(aliases_data)
        
        # Create a temporary file to write the CSV
        temp_file = os.path.join(app.config['UPLOAD_FOLDER'], 'aliases_export.csv')
        df.to_csv(temp_file, index=False)
        
        # Return the file as an attachment
        return send_file(temp_file, as_attachment=True, 
                         download_name='aliases.csv', 
                         mimetype='text/csv')
    except Exception as e:
        print(f"Error exporting aliases: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/api/aliases/import', methods=['POST'])
def import_aliases():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file and (file.filename.endswith('.csv') or file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        try:
            # Save the uploaded file
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            # Read the file based on its extension
            if file.filename.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path)
            
            # Check if required columns exist
            required_columns = ['Ingredient Name', 'Alias']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return jsonify({
                    'error': f'Missing required columns: {", ".join(missing_columns)}'
                }), 400
            
            # Process the aliases
            aliases_added = 0
            errors = 0
            
            for index, row in df.iterrows():
                ingredient_name = row['Ingredient Name']
                alias_text = row['Alias']
                
                # Skip rows with missing data
                if pd.isna(ingredient_name) or pd.isna(alias_text):
                    errors += 1
                    continue
                
                # Find the ingredient by name
                ingredient = Ingredient.query.filter(
                    Ingredient.name.ilike(f'{ingredient_name}')
                ).first()
                
                if not ingredient:
                    errors += 1
                    continue
                
                # Check if the alias already exists
                existing_alias = IngredientAlias.query.filter_by(
                    ingredient_id=ingredient.id,
                    alias=alias_text
                ).first()
                
                if not existing_alias:
                    # Add the new alias
                    new_alias = IngredientAlias(
                        alias=alias_text,
                        ingredient_id=ingredient.id
                    )
                    db.session.add(new_alias)
                    aliases_added += 1
            
            # Commit the changes
            db.session.commit()
            
            # Clean up the file
            if os.path.exists(file_path):
                os.remove(file_path)
            
            return jsonify({
                'message': f'Successfully imported {aliases_added} aliases',
                'aliases_added': aliases_added,
                'errors': errors
            })
        except Exception as e:
            db.session.rollback()
            print(f"Error importing aliases: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({"error": str(e)}), 500
    
    return jsonify({'error': 'Invalid file format. Please upload a CSV or Excel file.'}), 400

@app.route('/api/filter-options', methods=['GET'])
def get_filter_options():
    try:
        # Get unique brands
        brands_query = db.session.query(distinct(Formula.formula_brand)).filter(Formula.formula_brand != '').order_by(Formula.formula_brand)
        brands = [brand[0] for brand in brands_query.all()]
        
        # Get unique categories
        categories_query = db.session.query(distinct(Formula.sbu_category)).filter(Formula.sbu_category != '').order_by(Formula.sbu_category)
        categories = [category[0] for category in categories_query.all()]
        
        # Get unique lifecycle phases
        phases_query = db.session.query(distinct(Formula.lifecycle_phase)).filter(Formula.lifecycle_phase != '').order_by(Formula.lifecycle_phase)
        lifecycle_phases = [phase[0] for phase in phases_query.all()]
        
        # Get unique production sites (split by comma, strip whitespace, flatten)
        sites_query = db.session.query(Formula.production_sites).filter(Formula.production_sites != '').distinct()
        sites_set = set()
        for row in sites_query:
            if row[0]:
                for site in row[0].split(','):
                    site_clean = site.strip()
                    if site_clean:
                        sites_set.add(site_clean)
        production_sites = sorted(sites_set)
        
        return jsonify({
            'brands': brands,
            'categories': categories,
            'lifecyclePhases': lifecycle_phases,
            'productionSites': production_sites
        })
        
    except Exception as e:
        print(f"Error getting filter options: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/api/ingredients', methods=['GET'])
def get_ingredients():
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        search = request.args.get('search', '')
        
        # Base query joining with aliases to allow searching by alias
        query = Ingredient.query.outerjoin(IngredientAlias)
        
        if search:
            # Search by ingredient name OR alias
            query = query.filter(
                db.or_(
                    Ingredient.name.ilike(f'%{search}%'),
                    Ingredient.fing_item_number.ilike(f'%{search}%'),
                    IngredientAlias.alias.ilike(f'%{search}%')
                )
            )
        
        # Make the query distinct to avoid duplicates from the join
        query = query.distinct()
            
        # Apply pagination
        pagination = query.order_by(Ingredient.name).paginate(page=page, per_page=per_page, error_out=False)
        
        ingredients = []
        for ingredient in pagination.items:
            aliases = [alias.alias for alias in ingredient.aliases]
            ingredients.append({
                'id': ingredient.id,
                'name': ingredient.name,
                'fing_item_number': ingredient.fing_item_number,
                'description': ingredient.description,
                'aliases': aliases
            })
        
        return jsonify({
            'ingredients': ingredients,
            'pagination': {
                'total': pagination.total,
                'pages': pagination.pages,
                'current_page': page,
                'per_page': per_page,
                'has_next': pagination.has_next,
                'has_prev': pagination.has_prev
            }
        })
    except Exception as e:
        print(f"Error in get_ingredients: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/api/ingredients/<int:ingredient_id>/aliases', methods=['GET'])
def get_ingredient_aliases(ingredient_id):
    try:
        ingredient = Ingredient.query.get_or_404(ingredient_id)
        aliases = [{'id': alias.id, 'alias': alias.alias} for alias in ingredient.aliases]
        
        return jsonify({
            'ingredient_id': ingredient_id,
            'ingredient_name': ingredient.name,
            'aliases': aliases
        })
    except Exception as e:
        print(f"Error in get_ingredient_aliases: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/ingredients/<int:ingredient_id>/aliases', methods=['POST'])
def add_ingredient_alias(ingredient_id):
    try:
        ingredient = Ingredient.query.get_or_404(ingredient_id)
        data = request.json
        
        if not data or 'alias' not in data:
            return jsonify({"error": "Alias is required"}), 400
            
        alias_text = data['alias'].strip()
        if not alias_text:
            return jsonify({"error": "Alias cannot be empty"}), 400
            
        # Check if alias already exists for this ingredient
        existing_alias = IngredientAlias.query.filter_by(
            ingredient_id=ingredient_id, 
            alias=alias_text
        ).first()
        
        if existing_alias:
            return jsonify({"error": "This alias already exists for this ingredient"}), 400
            
        # Create new alias
        new_alias = IngredientAlias(
            alias=alias_text,
            ingredient_id=ingredient_id
        )
        
        db.session.add(new_alias)
        db.session.commit()
        
        return jsonify({
            'id': new_alias.id,
            'alias': new_alias.alias,
            'ingredient_id': ingredient_id,
            'message': 'Alias added successfully'
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error in add_ingredient_alias: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/aliases/<int:alias_id>', methods=['DELETE'])
def delete_alias(alias_id):
    try:
        alias = IngredientAlias.query.get_or_404(alias_id)
        ingredient_id = alias.ingredient_id
        
        db.session.delete(alias)
        db.session.commit()
        
        return jsonify({
            'message': 'Alias deleted successfully',
            'ingredient_id': ingredient_id
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error in delete_alias: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/upload-excel', methods=['POST'])
def upload_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        try:
            # Load the Excel file
            df = pd.read_excel(file_path)
            
            # Check if required columns exist
            required_columns = [
                'FING_ITEM_NUMBER', 'FING_DESCRIPTION', 
                'Ingredient Description Expanded', 'FING_QUANTITY', 'FING_UNIT', 
                'OBJECT_NUMBER', 'LIFECYCLE_PHASE', 'FORMULATION_NAME'
            ]
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return jsonify({
                    'error': f'Missing required columns: {", ".join(missing_columns)}'
                }), 400
            
            # Clear existing data
            db.session.query(FormulaIngredient).delete()
            db.session.query(Formula).delete()
            db.session.query(Ingredient).delete()
            db.session.commit()
            
            # Process the data
            ingredients_dict = {}
            formulas_dict = {}
            formula_ingredients_created = 0
            errors = 0
            
            # Process each row in the Excel file
            for index, row in df.iterrows():
                try:
                    # Process ingredient if it doesn't exist
                    fing_item_number = str(row['FING_ITEM_NUMBER'])
                    if fing_item_number not in ingredients_dict:
                        ingredient = Ingredient(
                            fing_item_number=fing_item_number,
                            name=str(row['FING_DESCRIPTION']),
                            description=str(row['FING_DESCRIPTION']),
                            description_expanded=str(row['Ingredient Description Expanded'])
                        )
                        db.session.add(ingredient)
                        db.session.flush()
                        ingredients_dict[fing_item_number] = ingredient.id
                    
                    # Process formula if it doesn't exist
                    object_number = str(row['OBJECT_NUMBER'])
                    if object_number not in formulas_dict:
                        formula = Formula(
                            object_number=object_number,
                            formulation_name=str(row['FORMULATION_NAME']),
                            lifecycle_phase=str(row['LIFECYCLE_PHASE']),
                            formula_brand=str(row['FORMULA_BRAND']) if 'FORMULA_BRAND' in row else '',
                            sbu_category=str(row['SBU_CATEGORY']) if 'SBU_CATEGORY' in row else '',
                            dossier_type=str(row['DOSSIER_TYPE']) if 'DOSSIER_TYPE' in row else '',
                            regulatory_comments=str(row['REGULATORY_COMMENTS']) if 'REGULATORY_COMMENTS' in row and not pd.isna(row['REGULATORY_COMMENTS']) else '',
                            general_comments=str(row['GENERAL_COMMENTS']) if 'GENERAL_COMMENTS' in row and not pd.isna(row['GENERAL_COMMENTS']) else '',
                            production_sites=str(row['PRODUCTION_SITES_AVAILABLE']) if 'PRODUCTION_SITES_AVAILABLE' in row and not pd.isna(row['PRODUCTION_SITES_AVAILABLE']) else '',
                            predecessor_formulation_number=str(row['PREDECESSORFORMULATIONNUMBER']) if 'PREDECESSORFORMULATIONNUMBER' in row and not pd.isna(row['PREDECESSORFORMULATIONNUMBER']) else '',
                            successor_formulation_number=str(row['SUCCESSORFORMULATIONNUMBER']) if 'SUCCESSORFORMULATIONNUMBER' in row and not pd.isna(row['SUCCESSORFORMULATIONNUMBER']) else ''
                        )
                        db.session.add(formula)
                        db.session.flush()
                        formulas_dict[object_number] = formula.id
                    
                    # Create formula ingredient relationship
                    # Handle NaN values in the amount field
                    amount = 0.0  # Default value
                    if 'FING_QUANTITY' in row:
                        if pd.notna(row['FING_QUANTITY']):
                            try:
                                # Try to convert to float
                                amount = float(row['FING_QUANTITY'])
                            except (ValueError, TypeError):
                                # If conversion fails, use the default value
                                if isinstance(row['FING_QUANTITY'], str) and row['FING_QUANTITY'].strip().upper() == 'Q.S.':
                                    # Special case for "Q.S." (Quantum Satis - as much as needed)
                                    amount = 0.0  # Use 0 as a placeholder
                                else:
                                    print(f"Row {index}: Invalid amount value: {row['FING_QUANTITY']}, using 0.0")
                    
                    unit = str(row['FING_UNIT']) if pd.notna(row['FING_UNIT']) else ''
                    
                    formula_ingredient = FormulaIngredient(
                        formula_id=formulas_dict[object_number],
                        ingredient_id=ingredients_dict[fing_item_number],
                        amount=amount,
                        unit=unit
                    )
                    db.session.add(formula_ingredient)
                    formula_ingredients_created += 1
                    
                except Exception as e:
                    errors += 1
                    print(f"Error processing row {index}: {str(e)}")
                    continue
                
                # Commit in batches
                if index % 1000 == 0:
                    try:
                        db.session.commit()
                    except Exception as e:
                        db.session.rollback()
                        print(f"Error committing batch at row {index}: {str(e)}")
            
            # Final commit
            db.session.commit()
            
            # Restore aliases from the latest backup
            print("Restoring aliases from latest backup...")
            latest_backup = get_latest_backup()
            if latest_backup:
                aliases_restored = restore_aliases_from_backup(latest_backup)
                print(f"Restored {aliases_restored} aliases from backup")
                
                # Include in the response
                return jsonify({
                    'success': True,
                    'message': 'Excel file imported successfully',
                    'ingredients_created': len(ingredients_dict),
                    'formulas_created': len(formulas_dict),
                    'formula_ingredients_created': formula_ingredients_created,
                    'aliases_restored': aliases_restored,
                    'errors': errors
                })
            else:
                return jsonify({
                    'success': True,
                    'message': 'Excel file imported successfully',
                    'ingredients_created': len(ingredients_dict),
                    'formulas_created': len(formulas_dict),
                    'formula_ingredients_created': formula_ingredients_created,
                    'aliases_restored': 0,
                    'errors': errors
                })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Error importing Excel file: {str(e)}'}), 500
        finally:
            # Clean up the uploaded file
            if os.path.exists(file_path):
                os.remove(file_path)
    
    return jsonify({'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls).'}), 400

@app.route('/api/database-status', methods=['GET'])
def database_status():
    try:
        ingredient_count = Ingredient.query.count()
        formula_count = Formula.query.count()
        formula_ingredient_count = FormulaIngredient.query.count()
        
        return jsonify({
            'success': True,
            'status': 'Database is available',
            'counts': {
                'ingredients': ingredient_count,
                'formulas': formula_count,
                'formula_ingredients': formula_ingredient_count
            },
            'has_data': ingredient_count > 0 and formula_count > 0
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'status': f'Error checking database: {str(e)}',
            'has_data': False
        })

@app.route('/api/initialize-database', methods=['POST'])
def initialize_database():
    try:
        success = initialize_database_from_excel()
        
        if success:
            # Get count of restored aliases
            aliases_count = IngredientAlias.query.count()
            
            return jsonify({
                'success': True,
                'message': 'Database initialized successfully from Excel file',
                'aliases_restored': aliases_count
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Failed to initialize database. Check server logs for details.'
            }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error initializing database: {str(e)}'
        }), 500

@app.route('/api/aliases/server-backup', methods=['POST'])
def create_server_backup():
    try:
        # Get all aliases with ingredient information
        aliases = []
        for alias in IngredientAlias.query.all():
            ingredient = Ingredient.query.get(alias.ingredient_id)
            if ingredient:
                aliases.append({
                    'ingredient_name': ingredient.name,
                    'ingredient_number': ingredient.fing_item_number,
                    'alias': alias.alias
                })
        
        # Create a timestamp for the backup filename
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f"aliases_backup_{timestamp}.json"
        backup_path = os.path.join(BACKUPS_DIR, backup_filename)
        
        # Write the backup to a file
        with open(backup_path, 'w') as f:
            json.dump(aliases, f, indent=2)
        
        return jsonify({
            'success': True,
            'message': 'Backup created successfully',
            'backup_id': timestamp,
            'filename': backup_filename
        })
    except Exception as e:
        print(f"Error creating backup: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/aliases/server-backups', methods=['GET'])
def list_server_backups():
    try:
        backups = []
        # List all JSON files in the backups directory
        for filename in os.listdir(BACKUPS_DIR):
            if filename.endswith('.json') and filename.startswith('aliases_backup_'):
                file_path = os.path.join(BACKUPS_DIR, filename)
                file_stats = os.stat(file_path)
                
                # Extract timestamp from filename
                timestamp = filename.replace('aliases_backup_', '').replace('.json', '')
                
                # Get the number of aliases in the backup
                try:
                    with open(file_path, 'r') as f:
                        aliases_count = len(json.load(f))
                except:
                    aliases_count = 0
                
                backups.append({
                    'id': timestamp,
                    'filename': filename,
                    'created_at': datetime.datetime.fromtimestamp(file_stats.st_mtime).isoformat(),
                    'size': file_stats.st_size,
                    'aliases_count': aliases_count
                })
        
        # Sort backups by creation time (newest first)
        backups.sort(key=lambda x: x['created_at'], reverse=True)
        
        return jsonify({
            'backups': backups
        })
    except Exception as e:
        print(f"Error listing backups: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/aliases/server-restore/<backup_id>', methods=['POST'])
def restore_server_backup(backup_id):
    try:
        # Find the backup file
        backup_filename = f"aliases_backup_{backup_id}.json"
        backup_path = os.path.join(BACKUPS_DIR, backup_filename)
        
        if not os.path.exists(backup_path):
            return jsonify({'error': 'Backup not found'}), 404
        
        # Load the backup data
        with open(backup_path, 'r') as f:
            aliases_data = json.load(f)
        
        aliases_restored = 0
        aliases_skipped = 0
        
        # Clear existing aliases if requested
        clear_existing = request.args.get('clear', 'false').lower() == 'true'
        if clear_existing:
            db.session.query(IngredientAlias).delete()
            db.session.commit()
        
        for alias_data in aliases_data:
            # Find the ingredient by name or number
            ingredient = Ingredient.query.filter(
                (Ingredient.name == alias_data['ingredient_name']) | 
                (Ingredient.fing_item_number == alias_data['ingredient_number'])
            ).first()
            
            if ingredient:
                # Check if this alias already exists for the ingredient
                existing_alias = IngredientAlias.query.filter_by(
                    ingredient_id=ingredient.id,
                    alias=alias_data['alias']
                ).first()
                
                if not existing_alias:
                    new_alias = IngredientAlias(
                        alias=alias_data['alias'],
                        ingredient_id=ingredient.id
                    )
                    db.session.add(new_alias)
                    aliases_restored += 1
                else:
                    aliases_skipped += 1
            else:
                aliases_skipped += 1
        
        # Commit the restored aliases
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Restored {aliases_restored} aliases successfully',
            'aliases_restored': aliases_restored,
            'aliases_skipped': aliases_skipped
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error restoring backup: {str(e)}")
        return jsonify({'error': f'Error restoring backup: {str(e)}'}), 500

@app.route('/api/aliases/server-backup/<backup_id>', methods=['DELETE'])
def delete_server_backup(backup_id):
    try:
        # Find the backup file
        backup_filename = f"aliases_backup_{backup_id}.json"
        backup_path = os.path.join(BACKUPS_DIR, backup_filename)
        
        if not os.path.exists(backup_path):
            return jsonify({'error': 'Backup not found'}), 404
        
        # Delete the backup file
        os.remove(backup_path)
        
        return jsonify({
            'success': True,
            'message': 'Backup deleted successfully'
        })
    except Exception as e:
        print(f"Error deleting backup: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/export-database', methods=['GET'])
def export_database():
    try:
        # Query all ingredients
        ingredients = []
        for ingredient in Ingredient.query.all():
            ingredients.append({
                'id': ingredient.id,
                'name': ingredient.name,
                'fing_item_number': ingredient.fing_item_number,
                'description': ingredient.description,
                'description_expanded': ingredient.description_expanded
            })
        
        # Query all formulas with their ingredients
        formulas = []
        for formula in Formula.query.all():
            formula_data = {
                'id': formula.id,
                'object_number': formula.object_number,
                'formulation_name': formula.formulation_name,
                'lifecycle_phase': formula.lifecycle_phase,
                'formula_brand': formula.formula_brand,
                'sbu_category': formula.sbu_category,
                'dossier_type': formula.dossier_type,
                'regulatory_comments': formula.regulatory_comments,
                'general_comments': formula.general_comments,
                'production_sites': formula.production_sites,
                'predecessor_formulation_number': formula.predecessor_formulation_number,
                'successor_formulation_number': formula.successor_formulation_number,
                'ingredients': []
            }
            
            for formula_ingredient in formula.ingredients:
                formula_data['ingredients'].append({
                    'ingredient_id': formula_ingredient.ingredient_id,
                    'amount': formula_ingredient.amount,
                    'unit': formula_ingredient.unit
                })
            
            formulas.append(formula_data)
        
        # Return the complete database
        return jsonify({
            'ingredients': ingredients,
            'formulas': formulas
        })
    except Exception as e:
        print(f"Error exporting database: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/api/upload-database', methods=['POST'])
def upload_database():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file and file.filename.endswith('.json'):
        try:
            # Read the JSON file
            json_data = json.load(file)
            
            # Validate the structure
            if 'ingredients' not in json_data or 'formulas' not in json_data:
                return jsonify({'error': 'Invalid JSON format: missing ingredients or formulas'}), 400
            
            # Back up aliases before clearing data
            aliases_backup = []
            for alias in IngredientAlias.query.all():
                ingredient = Ingredient.query.get(alias.ingredient_id)
                if ingredient:
                    aliases_backup.append({
                        'ingredient_name': ingredient.name,
                        'ingredient_number': ingredient.fing_item_number,
                        'alias': alias.alias
                    })
            
            # Create an automatic backup if there are aliases
            if aliases_backup:
                timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
                backup_filename = f"aliases_backup_{timestamp}.json"
                backup_path = os.path.join(BACKUPS_DIR, backup_filename)
                
                # Write the backup to a file
                with open(backup_path, 'w') as f:
                    json.dump(aliases_backup, f, indent=2)
            
            # Clear existing data
            db.session.query(FormulaIngredient).delete()
            db.session.query(Formula).delete()
            db.session.query(IngredientAlias).delete()
            db.session.query(Ingredient).delete()
            db.session.commit()
            
            # Create a mapping from file ingredient IDs to database IDs
            ingredient_id_mapping = {}
            
            # Import ingredients
            for ingredient_data in json_data['ingredients']:
                ingredient = Ingredient(
                    name=ingredient_data.get('name', ''),
                    fing_item_number=ingredient_data.get('fing_item_number', ''),
                    description=ingredient_data.get('description', ''),
                    description_expanded=ingredient_data.get('description_expanded', '')
                )
                db.session.add(ingredient)
                db.session.flush()  # Get the ID without committing
                ingredient_id_mapping[ingredient_data['id']] = ingredient.id
            
            # Import formulas
            for formula_data in json_data['formulas']:
                formula = Formula(
                    object_number=formula_data.get('object_number', ''),
                    formulation_name=formula_data.get('formulation_name', ''),
                    lifecycle_phase=formula_data.get('lifecycle_phase', ''),
                    formula_brand=formula_data.get('formula_brand', ''),
                    sbu_category=formula_data.get('sbu_category', ''),
                    dossier_type=formula_data.get('dossier_type', ''),
                    regulatory_comments=formula_data.get('regulatory_comments', ''),
                    general_comments=formula_data.get('general_comments', ''),
                    production_sites=formula_data.get('production_sites', ''),
                    predecessor_formulation_number=formula_data.get('predecessor_formulation_number', ''),
                    successor_formulation_number=formula_data.get('successor_formulation_number', '')
                )
                db.session.add(formula)
                db.session.flush()
                
                # Import formula ingredients
                for ingredient_data in formula_data.get('ingredients', []):
                    # Map the ingredient ID from the file to the database ID
                    if ingredient_data.get('ingredient_id') in ingredient_id_mapping:
                        db_ingredient_id = ingredient_id_mapping[ingredient_data['ingredient_id']]
                        
                        formula_ingredient = FormulaIngredient(
                            formula_id=formula.id,
                            ingredient_id=db_ingredient_id,
                            amount=ingredient_data.get('amount', 0.0),
                            unit=ingredient_data.get('unit', '')
                        )
                        db.session.add(formula_ingredient)
            
            # Commit all changes
            db.session.commit()
            
            # Restore aliases from backup
            aliases_restored = 0
            for alias_data in aliases_backup:
                # Find the ingredient by name or number
                ingredient = Ingredient.query.filter(
                    (Ingredient.name == alias_data['ingredient_name']) | 
                    (Ingredient.fing_item_number == alias_data['ingredient_number'])
                ).first()
                
                if ingredient:
                    # Check if this alias already exists for the ingredient
                    existing_alias = IngredientAlias.query.filter_by(
                        ingredient_id=ingredient.id,
                        alias=alias_data['alias']
                    ).first()
                    
                    if not existing_alias:
                        new_alias = IngredientAlias(
                            alias=alias_data['alias'],
                            ingredient_id=ingredient.id
                        )
                        db.session.add(new_alias)
                        aliases_restored += 1
            
            if aliases_restored > 0:
                db.session.commit()
            
            return jsonify({
                'message': 'Database imported successfully',
                'ingredients_count': len(json_data['ingredients']),
                'formulas_count': len(json_data['formulas']),
                'aliases_restored': aliases_restored
            })
        
        except json.JSONDecodeError:
            return jsonify({'error': 'Invalid JSON file'}), 400
        except Exception as e:
            db.session.rollback()
            print(f"Error importing database: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({"error": str(e)}), 500
    
    return jsonify({'error': 'Invalid file format. Please upload a JSON file.'}), 400

@app.route('/api/formulas/export', methods=['GET'])
def export_formulas_excel():
    try:
        # Get filter parameters
        brand = request.args.get('brand', '')
        category = request.args.get('category', '')
        lifecycle_phase = request.args.get('lifecycle_phase', '')

        # Build query with filters (reuse logic from get_formulas)
        query = Formula.query
        if brand:
            query = query.filter(Formula.formula_brand == brand)
        if category:
            query = query.filter(Formula.sbu_category == category)
        if lifecycle_phase:
            query = query.filter(Formula.lifecycle_phase == lifecycle_phase)

        formulas = query.all()

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Formulas'
        # Header
        headers = ['Object Number', 'Formulation Name', 'Lifecycle Phase', 'Brand', 'Category']
        ws.append(headers)
        # Data rows
        for formula in formulas:
            ws.append([
                formula.object_number,
                formula.formulation_name,
                formula.lifecycle_phase,
                formula.formula_brand,
                formula.sbu_category
            ])
        # Format as table
        last_row = ws.max_row
        last_col = ws.max_column
        table_ref = f"A1:{chr(64+last_col)}{last_row}"
        table = Table(displayName="FormulaTable", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
        # Set hard-coded column widths
        col_widths = [18, 32, 18, 18, 18]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        # Generate unique filename: FIND list export DDMMYY_N.xlsx
        today_str = datetime.datetime.now().strftime('%d%m%y')
        counter_file = os.path.join(BACKUPS_DIR, f'export_counter_{today_str}.txt')
        if os.path.exists(counter_file):
            with open(counter_file, 'r') as f:
                try:
                    counter = int(f.read().strip()) + 1
                except:
                    counter = 1
        else:
            counter = 1
        with open(counter_file, 'w') as f:
            f.write(str(counter))
        filename = f"FIND list export {today_str}_{counter}.xlsx"
        # Save to in-memory buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        # Send as file
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"Error exporting formulas to Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/formulas/<object_number>/pdf', methods=['GET'])
def export_formula_pdf(object_number):
    formula = Formula.query.filter_by(object_number=object_number).first_or_404()
    # Prepare data for template
    formula_data = {
        'object_number': formula.object_number,
        'formulation_name': formula.formulation_name,
        'lifecycle_phase': formula.lifecycle_phase,
        'formula_brand': formula.formula_brand,
        'sbu_category': formula.sbu_category,
        'dossier_type': formula.dossier_type,
        'regulatory_comments': formula.regulatory_comments,
        'general_comments': formula.general_comments,
        'production_sites': formula.production_sites,
        'predecessor_formulation_number': formula.predecessor_formulation_number,
        'successor_formulation_number': formula.successor_formulation_number,
        'ingredients': []
    }
    for formula_ingredient in formula.ingredients:
        ingredient = Ingredient.query.get(formula_ingredient.ingredient_id)
        formula_data['ingredients'].append({
            'name': ingredient.name,
            'fing_item_number': ingredient.fing_item_number,
            'amount': formula_ingredient.amount,
            'unit': formula_ingredient.unit
        })
    # Render HTML template
    html = render_template('formula_pdf.html', formula=formula_data)
    # Convert HTML to PDF
    pdf = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=pdf)
    if pisa_status.err:
        return jsonify({'error': 'PDF generation failed'}), 500
    pdf.seek(0)
    filename = f"Formula_{formula.object_number}.pdf"
    return send_file(pdf, as_attachment=True, download_name=filename, mimetype='application/pdf')

if __name__ == '__main__':
    with app.app_context():
        # Drop and recreate all tables
        print("Dropping all tables...")
        db.drop_all()
        print("Creating all tables...")
        db.create_all()
        
        # Try to initialize from Excel
        print("Attempting to initialize database from Excel file...")
        success = initialize_database_from_excel()
        
        if success:
            print("Database successfully initialized from Excel file!")
        else:
            print("Failed to initialize database from Excel file.")

    debug_mode = os.environ.get('FLASK_DEBUG', 'True').lower() == 'true'
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=debug_mode, host='0.0.0.0', port=port)